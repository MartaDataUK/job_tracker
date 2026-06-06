import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_scraped_description(text):
    """
    Surgically strips away scraping artifacts, header navigation blocks,
    and character encoding visual glitches without discarding any rows.
    """
    if not text or pd.isna(text):
        return ""

    text = str(text)

    junk_patterns = [
        r"What\?\s*Where\?\s*Search\s*Advanced",
        r"â\x9d®\s*back to last search",
        r"Receive similar jobs by email[\s\S]*?Create alert",
        r"By creating an alert, you agree to our[\s\S]*",
        r"No thanks, take me to the job",
        r"Stats for this job[\s\S]*",
        r"Popular searches[\s\S]*",
        r"Similar jobs[\s\S]*"
    ]
    for pattern in junk_patterns:
        text = re.sub(pattern, "", text, flags=re.IGNORECASE)

    text = text.replace("â€™", "'").replace("â€“", "–").replace("Â£", "£").replace("â\x9d®", "").replace("Â", "")
    text = re.sub(r'\n\s*\n', '\n\n', text)
    text = re.sub(r'[ \t]+', ' ', text)

    return text.strip()

def calculate_job_score(title, desc, location, salary_period=None):
    score = 0
    title_lower = str(title).lower()
    desc_lower = str(desc).lower()
    location_lower = str(location).lower()
    salary_period = str(salary_period).lower() if salary_period is not None and not pd.isna(salary_period) else ""

    # Job title scoring
    if "analytics engineer" in title_lower:
        score += 30
    elif "data engineer" in title_lower or "product developer" in title_lower:
        score += 15
    elif "data scientist" in title_lower:
        score += 20
    elif "data analyst" in title_lower:
        score += 25

    is_truncated = desc_lower.endswith("…") or "â€¦" in desc_lower or len(desc_lower) < 500
    if is_truncated and "analytics engineer" in title_lower:
        score += 15

    # Proper technical keywords
    if re.search(r'\bdbt\b', desc_lower) or "dbt" in title_lower:
        score += 10
    if "bigquery" in desc_lower or re.search(r'\bgcp\b', desc_lower) or "bq" in title_lower:
        score += 15
    if re.search(r'\bpython\b', desc_lower):
        score += 5

    if any(word in desc_lower for word in ["git", "version control", "code review", "ci/cd"]):
        score += 5
    if any(word in desc_lower for word in ["data quality", "testing", "assertion", "lineage"]):
        score += 10
    if "single source of truth" in desc_lower or "semantic layer" in desc_lower:
        score += 5
    if "autonomy" in desc_lower or "self-starter" in desc_lower or "roadmap" in desc_lower or "independent" in desc_lower:
        score += 5

    # Industry relevance signals
    if any(word in desc_lower or word in title_lower for word in ["telecom", "broadband", "network", "satellite", "aviation", "connectivity"]):
        score += 15
    elif any(word in desc_lower for word in ["ecommerce", "e-commerce", "retail", "inventory", "streaming", "media"]):
        score += 12
    elif any(word in desc_lower for word in ["insurance", "underwriting", "broker", "policy lifecycle"]):
        score += 10

    # 🏴󠁧󠁢󠁳󠁣󠁴󠁿 Location Definition For Scoring
    is_scotland = any(city in location_lower for city in ["edinburgh", "glasgow", "livingston", "dundee", "aberdeen", "scotland"])
    is_fully_remote = "fully remote" in desc_lower or "100% remote" in desc_lower or "fully remote" in location_lower

    # Core Priority: Maximum points for local proximity regardless of setup
    if is_scotland:
        score += 25
    # Secondary Priority: Strong baseline boost for out-of-region elite remote roles
    elif is_fully_remote:
        score += 35

    is_relevant_data_role = (
        "analytics" in title_lower or
        "data" in title_lower or
        "product developer" in title_lower or
        "dbt" in desc_lower or
        "bigquery" in desc_lower
        or re.search(r'\bpython\b', desc_lower)
    )

    if is_relevant_data_role:
        has_contract_signals = any(term in desc_lower for term in ["inside ir35", "umbrella", "ir35 inside", "day rate", "contractor", "interim"]) or salary_period == "daily"
        # Contract and IR35 intent detection
        if has_contract_signals:
            if any(term in desc_lower or term in title_lower for term in ["inside ir35", "umbrella", "ir35 inside"]):
                score += 35
            if any(term in desc_lower for term in ["day rate", "contractor", "interim", "initial 6"]):
                score += 30
            if salary_period == "daily":
                score += 15
        else:
            # 🌍 COUNTER-WEIGHT: Reward premium permanent structures so they don't get buried by contract points
            is_high_tier_perm = any(phrase in desc_lower for phrase in ["£70,000", "£75,000", "£80,000", "£85,000", "£90,000", "70k", "75k", "80k", "85k", "90k"])
            if is_high_tier_perm or any(term in desc_lower for term in ["permanent position", "bonus scheme", "private dental", "annual bonus"]):
                score += 30
        legacy_or_wrong_cloud_stack = [
        "hadoop", "spark", "scala", "athena", "glue", "kinesis", "lambda", "service desk"]

        if any(bad_tech in desc_lower for bad_tech in legacy_or_wrong_cloud_stack):
            score -= 20
    return score

def compile_job_tracker_pipeline():
    excel_output = "UK_Data_Engineering_Job_Tracker.xlsx"

    existing_todo = pd.DataFrame()
    existing_applied = pd.DataFrame()
    existing_progressed = pd.DataFrame()
    tracked_links = set()

    # 🛡️ DEEP HYPERLINK LOADER: Extract actual target URLs hidden behind 'Apply' cells
    if os.path.exists(excel_output):
        try:
            # Read metadata strings cleanly using pandas
            existing_todo = pd.read_excel(excel_output, sheet_name="📥 Jobs To Do")
            existing_applied = pd.read_excel(excel_output, sheet_name="🚀 Applied")
            existing_progressed = pd.read_excel(excel_output, sheet_name="📈 Progressed")

            # Deep parse actual URL paths using openpyxl directly
            wb_loader = load_workbook(excel_output, read_only=False, data_only=False)
            for sheet_name in ["📥 Jobs To Do", "🚀 Applied", "📈 Progressed"]:
                if sheet_name in wb_loader.sheetnames:
                    ws_load = wb_loader[sheet_name]
                    # Find out which column index corresponds to job_link
                    headers = [cell.value for cell in ws_load[1]]
                    if "job_link" in headers:
                        link_col_idx = headers.index("job_link") + 1
                        for row in ws_load.iter_rows(min_row=2, max_col=link_col_idx, min_col=link_col_idx):
                            cell = row[0]
                            if cell.hyperlink and cell.hyperlink.target:
                                tracked_links.add(str(cell.hyperlink.target).strip())
                            elif cell.value and str(cell.value).startswith("http"):
                                tracked_links.add(str(cell.value).strip())
            print(f"📖 Loaded existing tracker progress. Found {len(tracked_links)} distinct links tracked so far.")
        except Exception as e:
            print(f"ℹ️ Note: Couldn't parse existing tracking sheets ({e}). Starting clean.")

    source_files = {
        "Linkedin": "linkedin_and_jobserve_jobs.csv",
        "Adzuna": "adzuna_jobs.csv"
    }

    unified_columns = [
        "job_link", "platform", "job_title", "location", "job_employment_type",
        "job_maturity", "job_min_salary", "job_max_salary", "job_salary_period",
        "job_description", "employer_name", "employer_rating", "job_posted_date"
    ]

    column_order = ["job_posted_date", "Score", "Status", "job_title", "employer_name", "location", "platform", "job_link", "job_description"]
    applied_columns = column_order + ["Date Applied", "Application Method", "Notes"]
    progressed_columns = applied_columns + ["Next Step", "Interview Date", "Salary Offer"]

    # Ingest New CSV rows
    compiled_dfs = []
    for platform, filename in source_files.items():
        if os.path.exists(filename):
            try:
                temp_df = pd.read_csv(filename)
                if temp_df.empty:
                    print(f"ℹ️ {filename} is empty. Skipping...")
                    continue
                for col in unified_columns:
                    if col not in temp_df.columns:
                        temp_df[col] = None
                temp_df = temp_df[unified_columns]
                temp_df["platform"] = platform
                compiled_dfs.append(temp_df)
                print(f"✅ Loaded rows from {platform} ({filename})")
            except Exception as e:
                print(f"⚠️ Error reading source {filename}: {e}")
        else:
            print(f"ℹ️ {filename} not found yet. Skipping safely...")

    if not compiled_dfs:
        print("❌ Error: No new data sources found to compile.")
        return

    master_df = pd.concat(compiled_dfs, ignore_index=True)
    master_df = master_df.drop_duplicates(subset=["job_link"], keep="first")

    # ✂️ Drop incoming rows before processing if the URL already exists anywhere in Excel
    if tracked_links:
        master_df = master_df[~master_df["job_link"].astype(str).str.strip().isin(tracked_links)]

    print("🧹 Running parsing layout artifact cleaner across core dataset...")
    master_df["job_description"] = master_df["job_description"].apply(clean_scraped_description)

    filtered_rows = []

    # Title exclusions
    banned_title_keywords = [
        "algorithm developer", "ai engineer", "deployment engineer",
        "machine learning", "ml engineer", "weapon", "defense", "defence",
        "computer vision", "deep learning", "devops", "platform engineer"
    ]

    # Description exclusions for specialized ecosystems/tools you don't use
    banned_desc_keywords = [
        "pega", "pega cdh", "salesforce developer", "sap consultant",
        "dynamics 365", "service desk manager", "servicenow",
        "alteryx", "power apps", "power automate", "low-code"
    ]

    for _, row in master_df.iterrows():
        title_lower = str(row["job_title"]).lower()
        loc_lower = str(row["location"]).lower()
        desc_lower = str(row["job_description"]).lower()

        if any(banned_word in title_lower for banned_word in banned_title_keywords):
            continue

        if any(banned_skill in desc_lower for banned_skill in banned_desc_keywords):
            continue

        is_scotland = any(city in loc_lower for city in ["edinburgh", "glasgow", "livingston", "dundee", "aberdeen", "scotland"])
        is_hub_city = any(city in loc_lower for city in ["london", "manchester"])

        has_remote_metadata = "remote" in loc_lower or "remote" in title_lower
        has_remote_text = "100% remote" in desc_lower or "fully remote" in desc_lower or "work from home" in desc_lower

        one_day_patterns = [
            r"(1|one)\s*day\s*(a|per|\/)\s*week",
            r"once\s*(a|per)\s*week",
            r"1\s*day\s*in\s*(the\s*)?office",
            r"one\s*day\s*in\s*(the\s*)?office",
            r"hybrid\s*\(?1\s*day",
            r"flexible\s*.*\b1\s*day\b"
        ]
        is_exactly_one_day = any(re.search(pattern, desc_lower) for pattern in one_day_patterns)

        is_heavy_hybrid_trap = any(phrase in desc_lower for phrase in [
            "2 days", "3 days", "4 days", "two days", "three days", "four days",
            "minimum 2", "minimum 3", "minimum 4", "min 2", "min 3", "min 4",
            "twice a week", "3-4 days", "2-3 days", "in-office 4 days", "isn't for you if you want fully remote"
        ])

        is_valid_remote = (has_remote_metadata or has_remote_text) and not is_heavy_hybrid_trap
        is_valid_one_day_hub = is_hub_city and is_exactly_one_day and not is_heavy_hybrid_trap

        if is_scotland:
            filtered_rows.append(row)
        elif not is_scotland and (is_valid_remote or is_valid_one_day_hub):
            filtered_rows.append(row)
        else:
            continue

    # 🤝 SYSTEM MERGE PROCESSOR
    if filtered_rows:
        new_scraped_df = pd.DataFrame(filtered_rows).reset_index(drop=True)
        new_scraped_df["Score"] = new_scraped_df.apply(
            lambda r: calculate_job_score(r["job_title"], r["job_description"], r["location"]), axis=1
        )
        new_scraped_df["Status"] = "To Do"
        new_scraped_df["job_posted_date"] = new_scraped_df["job_posted_date"].fillna("04/06/2026")
        new_scraped_df.loc[new_scraped_df["job_posted_date"].astype(str).str.strip() == "", "job_posted_date"] = "04/06/2026"

        total_scraped_items = len(new_scraped_df)
        new_todo_df = new_scraped_df[column_order].copy()

        if not existing_todo.empty:
            # Overwrite the 'job_link' column back into existing dataframe safely
            if "job_link" in existing_todo.columns:
                existing_todo = existing_todo.reindex(columns=column_order)
            todo_df = pd.concat([existing_todo, new_todo_df], ignore_index=True)
        else:
            todo_df = new_todo_df

        if not todo_df.empty and "job_link" in todo_df.columns:
            todo_df = todo_df.drop_duplicates(subset=["job_link"], keep="first")

        todo_df = todo_df.sort_values(by="Score", ascending=False).reset_index(drop=True)
        new_unique_items_count = len(todo_df) - len(existing_todo)

        print(f"\n✨ --- Staging Merge Diagnostics ---")
        print(f"📥 Total incoming jobs processed through criteria filters: {total_scraped_items}")
        print(f"🚀 Successfully appended to To Do sheet: +{max(0, new_unique_items_count)} brand new leads!")
        print(f"------------------------------------\n")
    else:
        # 💡 Fallback if no fresh scrapers generated new targets
        print("\nℹ️ No brand new items found via live scraper engines today.")
        todo_df = existing_todo if not existing_todo.empty else pd.DataFrame(columns=column_order)
        if not todo_df.empty:
            todo_df = todo_df.reindex(columns=column_order)

    # Re-assign progress trackers safely back into scope
    applied_df = existing_applied if not existing_applied.empty else pd.DataFrame(columns=applied_columns)
    progressed_df = existing_progressed if not existing_progressed.empty else pd.DataFrame(columns=progressed_columns)

    # Ensure metadata layers maintain exact structure mappings
    applied_df = applied_df.reindex(columns=applied_columns)
    progressed_df = progressed_df.reindex(columns=progressed_columns)

    # Workbook compilation setup
    wb = Workbook()
    wb.remove(wb.active)

    sheets_config = [
        ("📥 Jobs To Do", todo_df),
        ("🚀 Applied", applied_df),
        ("📈 Progressed", progressed_df)
    ]

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, bold=False, color="2C3E50")
    link_font = Font(name="Segoe UI", size=10, bold=False, color="2980B9", underline="single")
    score_font = Font(name="Segoe UI", size=10, bold=True, color="27AE60")

    thin_border = Border(
        left=Side(style='thin', color='E5E7E9'), right=Side(style='thin', color='E5E7E9'),
        top=Side(style='thin', color='E5E7E9'), bottom=Side(style='thin', color='E5E7E9')
    )

    for sheet_name, df_source in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        for r_idx, row in enumerate(dataframe_to_rows(df_source, index=False, header=True), start=1):
            ws.append(row)
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)

                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[r_idx].height = 26
                else:
                    cell.font = data_font
                    cell.border = thin_border
                    ws.row_dimensions[r_idx].height = 20
                    cell.fill = zebra_fill if r_idx % 2 == 0 else white_fill

                    col_name = df_source.columns[c_idx - 1]
                    if col_name in ["Status", "Score", "Date Applied", "Interview Date", "job_posted_date"]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    if col_name == "Score":
                        cell.font = score_font
                    elif col_name == "job_link" and cell.value:
                        link_val = str(cell.value)
                        if link_val.startswith("http") or link_val == "Apply":
                            # Pull target url fallback from dataframe context if openpyxl text is parsed
                            raw_target = df_source.iloc[r_idx - 2]["job_link"] if link_val == "Apply" else link_val

                            # 🎯 FIX: Keep the actual raw URL visible instead of "Apply"
                            cell.value = raw_target
                            cell.hyperlink = raw_target
                            cell.font = link_font
                            # Left-align the raw link so it reads cleanly in Excel
                            cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or '')
                if cell.hyperlink:
                    val_str = "Apply"
                if len(val_str) > max_len:
                    max_len = len(val_str)
                if col_letter == 'I': # Column I is job_description
                    max_len = min(max_len, 45)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

    wb.save(excel_output)
    print(f"🏁 Done! Master tracker dashboard successfully compiled at: '{excel_output}'")

if __name__ == "__main__":
    compile_job_tracker_pipeline()